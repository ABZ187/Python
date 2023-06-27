import openpyxl as xl
import openpyxl.utils as xlu
import openpyxl.styles as xls
import pywhatkit as wt

wl = xl.load_workbook("C:/Users/Admin/Desktop/Payments.xlsx")
ws = wl.active
# print(ws['B2'].value)
# ws['B2'] = "Test"
# print(ws['B2'].value)
# wl.save('C:/Users/Admin/Desktop/Payments.xlsx')
# wl.create_sheet("test")
# wl.save('C:/Users/Admin/Desktop/Payments.xlsx')

# print(wl.sheetnames)
# wk1 = xl.Workbook()
# ws1 = wk1.active
# wk1.title = 'workbook1'
# wk1.save('C:/Users/Admin/Desktop/workbook1.xlsx')
# ws1.append(["Jan","Feb","March"])
# wk1.save('C:/Users/Admin/Desktop/workbook1.xlsx')

# for row in range(2, 5):
#     for col in range(1, 5):
#         char = xlu.get_column_letter(col)
#         print(ws[char + str(row)].value)
wb = xl.load_workbook('C:/Users/Admin/Desktop/workbook1.xlsx')
ws2 = wb.active
# ws2.merge_cells('A2:D2')
# ws2.unmerge_cells('A2:D2')
# ws2.insert_rows(3)
# ws2.insert_cols(2)
# ws2.delete_cols(2)
# ws2.delete_rows(3)
# ws2.move_range('A2:B7', rows=3, cols=2)
x = '000000'
# Fill_pattern = xls.PatternFill(fill_type='solid', fgColor=x)
# ws2['B4'].fill=Fill_pattern
# color_hex=ws2['J8'].fill.start_color.index
# print(color_hex)

mon_row = str(input('Enter the row of month'))
for row in range(1, 13):
    y = mon_row + str(row)
    color_hex = ws2[y].fill.start_color.index
    a=str(color_hex)
    print(a)
    if a == '00000000':
        ws2[y] = 'message'
    else:
        ws2[y] = 'ho'

wb.save('C:/Users/Admin/Desktop/workbook1.xlsx')
